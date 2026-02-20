import pandas as pd
import requests
import re
import time
import json
from collections import defaultdict, Counter


# =========================
# CONFIG
# =========================
PARTIAL_DECKLIST_FILE = "mazos_a_medias.txt"  # Tu archivo con mazos parciales
INVENTORY_FILE = "inventario.csv"
OUTPUT_DIR = "mazos_completados"
RATE_LIMIT = 0.05
MIN_SYNERGY_SCORE = 0.10  # Reducido de 0.15 (más permisivo)
MAX_SUGGESTIONS = 50       # Aumentado de 30 (más sugerencias)
KEY_CARD_INCLUSION_MIN = 0.40  # 40% mínimo para "carta clave"
KEY_CARD_PRICE_MAX = 2.00  # $2 USD máximo
MIN_SCORE_THRESHOLD = 5    # Score mínimo para considerar
CHECK_COLOR_IDENTITY = True  # CRÍTICO: Verificar colores
COLOR_CHECK_RATE_LIMIT = 0.1  # Rate limit para Scryfall API


# =========================
# SESSION
# =========================
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

def create_session():
    session = requests.Session()
    retry = Retry(total=3, backoff_factor=1, status_forcelist=[429, 500, 502, 503, 504])
    adapter = HTTPAdapter(max_retries=retry)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    return session

SESSION = create_session()


# =========================
# 1. CARGAR INVENTARIO
# =========================
def load_inventory():
    """Carga y procesa el inventario"""
    inv = pd.read_csv(INVENTORY_FILE)
    inv["name_lower"] = inv["name"].str.lower()
    
    # Crear mapa de cartas disponibles
    inv_map = {}
    for _, row in inv.iterrows():
        name_lower = row["name_lower"]
        if name_lower not in inv_map:
            inv_map[name_lower] = {
                "name": row["name"],
                "quantity": 0,
                "collections": []
            }
        inv_map[name_lower]["quantity"] += row["quantity"]
        inv_map[name_lower]["collections"].append(str(row["source"]))
    
    return inv_map


# =========================
# 2. LEER MAZOS PARCIALES
# =========================
def parse_partial_decklists(filepath):
    """
    Lee archivo con mazos parciales en formato:
    
    # Comandante: Atraxa, Praetors' Voice
    1 Sol Ring
    1 Arcane Signet
    ...
    
    # Comandante: Muldrotha, the Gravetide
    1 Eternal Witness
    ...
    """
    decks = []
    current_deck = None
    
    with open(filepath, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            
            if line.startswith("# Comandante:") or line.startswith("#Comandante:"):
                if current_deck:
                    decks.append(current_deck)
                
                commander = line.split(":", 1)[1].strip()
                current_deck = {
                    "commander": commander,
                    "cards": []
                }
            
            elif line and not line.startswith("#") and current_deck:
                # Parsear línea tipo "1 Card Name" o "Card Name"
                parts = line.split(None, 1)
                if len(parts) == 2 and parts[0].isdigit():
                    card_name = parts[1]
                elif len(parts) == 1:
                    card_name = parts[0]
                else:
                    continue
                
                # Limpiar tags y marcadores
                card_name = re.sub(r'#\w+', '', card_name)  # Remover tags
                card_name = re.sub(r'[⭐🔥\[\]]', '', card_name)  # Remover marcadores
                card_name = card_name.strip()
                
                if card_name:
                    current_deck["cards"].append(card_name)
        
        if current_deck:
            decks.append(current_deck)
    
    return decks


# =========================
# 3. OBTENER IDENTIDAD DE COLORES (MEJORADO)
# =========================
def get_commander_colors(commander_name):
    """
    Obtiene la identidad de colores del comandante desde EDHREC
    Intenta múltiples métodos para garantizar resultado
    """
    url = f"https://json.edhrec.com/pages/commanders/{slug(commander_name)}.json"
    
    try:
        r = SESSION.get(url, timeout=10)
        if r.status_code != 200:
            print(f"      ⚠️  No se pudo obtener página del comandante (código {r.status_code})")
            return guess_colors_from_name(commander_name)
        
        data = r.json()
        
        # Método 1: Buscar "coloridentity" directamente
        def find_color_identity(obj, depth=0):
            if depth > 10:  # Evitar recursión infinita
                return None
            
            if isinstance(obj, dict):
                # Buscar coloridentity
                if "coloridentity" in obj:
                    return obj["coloridentity"]
                
                # Buscar colors (alternativo)
                if "colors" in obj and isinstance(obj["colors"], list):
                    return obj["colors"]
                
                # Buscar recursivamente
                for v in obj.values():
                    result = find_color_identity(v, depth + 1)
                    if result:
                        return result
            
            elif isinstance(obj, list):
                for item in obj:
                    result = find_color_identity(item, depth + 1)
                    if result:
                        return result
            
            return None
        
        colors = find_color_identity(data)
        
        # Método 2: Buscar en container.json_dict si existe
        if not colors and "container" in data:
            container = data["container"]
            if "json_dict" in container:
                json_dict = container["json_dict"]
                if "card" in json_dict:
                    card_info = json_dict["card"]
                    if "color_identity" in card_info:
                        colors = card_info["color_identity"]
        
        # Validar resultado
        if colors and isinstance(colors, list):
            # Normalizar: convertir a uppercase
            colors = [c.upper() if isinstance(c, str) else c for c in colors]
            print(f"      ✅ Colores detectados: {colors}")
            return colors
        
        print(f"      ⚠️  No se encontró identidad de colores en datos EDHREC")
        return guess_colors_from_name(commander_name)
        
    except Exception as e:
        print(f"      ❌ Error obteniendo colores: {e}")
        return guess_colors_from_name(commander_name)


def guess_colors_from_name(commander_name):
    """
    Intenta adivinar colores basándose en palabras clave en el nombre
    Último recurso si EDHREC no provee la info
    """
    name_lower = commander_name.lower()
    
    # Patrones comunes
    if "atraxa" in name_lower:
        return ["W", "U", "B", "G"]  # WUBG
    elif "muldrotha" in name_lower:
        return ["U", "B", "G"]  # UBG
    elif "korvold" in name_lower:
        return ["B", "R", "G"]  # BRG
    elif "golos" in name_lower:
        return ["W", "U", "B", "R", "G"]  # WUBRG
    elif "chulane" in name_lower:
        return ["G", "W", "U"]  # GWU
    
    # Si no lo conocemos, asumir 5 colores (permisivo)
    print(f"      ⚠️  Colores desconocidos para '{commander_name}', asumiendo 5 colores")
    return ["W", "U", "B", "R", "G"]


def card_is_legal_in_colors(card_name, commander_colors, inventory):
    """
    Verifica si una carta es legal en la identidad de colores del comandante
    Usa Scryfall API para obtener la identidad de la carta
    """
    # Si no tenemos colores del comandante, aceptar todo
    if not commander_colors:
        return True
    
    # Normalizar colores del comandante
    commander_colors_set = set(c.upper() for c in commander_colors)
    
    # Buscar en caché primero
    card_lower = card_name.lower()
    
    # Consultar Scryfall para obtener color identity
    try:
        import urllib.parse
        encoded = urllib.parse.quote(card_name)
        url = f"https://api.scryfall.com/cards/named?exact={encoded}"
        
        r = SESSION.get(url, timeout=5)
        if r.status_code == 200:
            data = r.json()
            card_colors = data.get("color_identity", [])
            
            # Verificar si la carta es legal
            card_colors_set = set(card_colors)
            is_legal = card_colors_set.issubset(commander_colors_set)
            
            return is_legal
        
        # Si falla Scryfall, asumir legal (permisivo)
        return True
        
    except Exception as e:
        # En caso de error, asumir legal
        return True


# =========================
# 4. SLUG
# =========================
def slug(s: str) -> str:
    import unicodedata
    s = unicodedata.normalize('NFD', s)
    s = ''.join(char for char in s if unicodedata.category(char) != 'Mn')
    s = s.lower()
    s = s.replace("'","").replace("'","").replace("â€™","")
    s = s.replace(",","").replace(":","").replace(".","")
    s = re.sub(r"[^a-z0-9]+","-",s)
    return s.strip("-")


# =========================
# 5. OBTENER AVERAGE DECK Y BUDGET + SCRYFALL
# =========================
def get_average_and_budget_deck(commander_name):
    """
    Obtiene cartas del average deck y budget deck
    Retorna dict con nombre_carta: {inclusion, num_decks, scryfall, ...}
    """
    url = f"https://json.edhrec.com/pages/average-decks/{slug(commander_name)}.json"
    
    try:
        r = SESSION.get(url, timeout=10)
        if r.status_code != 200:
            return {}
        
        data = r.json()
    except:
        return {}
    
    cards_info = {}
    
    def scan(obj, in_budget_section=False):
        if isinstance(obj, dict):
            # Detectar sección budget
            is_budget = "budget" in str(obj.get("tag", "")).lower() or in_budget_section
            
            if "cardviews" in obj:
                for card in obj["cardviews"]:
                    name = card.get("name")
                    if not name:
                        continue
                    
                    name_lower = name.lower()
                    
                    if name_lower not in cards_info:
                        cards_info[name_lower] = {
                            "name": name,
                            "inclusion": 0,
                            "num_decks": 0,
                            "is_budget": is_budget,
                            "synergy": 0,
                            "price": card.get("price", 0),
                            "scryfall": card.get("scryfall_uri", "")
                        }
                    
                    # Actualizar información
                    inclusion = card.get("inclusion", 0)
                    num_decks = card.get("num_decks", 0)
                    scryfall = card.get("scryfall_uri", "")
                    
                    if inclusion > cards_info[name_lower]["inclusion"]:
                        cards_info[name_lower]["inclusion"] = inclusion
                    if num_decks > cards_info[name_lower]["num_decks"]:
                        cards_info[name_lower]["num_decks"] = num_decks
                    if scryfall and not cards_info[name_lower]["scryfall"]:
                        cards_info[name_lower]["scryfall"] = scryfall
                    
                    cards_info[name_lower]["is_budget"] = cards_info[name_lower]["is_budget"] or is_budget
            
            for v in obj.values():
                scan(v, is_budget)
        
        elif isinstance(obj, list):
            for item in obj:
                scan(item, in_budget_section)
    
    scan(data)
    return cards_info


# =========================
# 6. OBTENER SINERGIAS CARTA A CARTA + SCRYFALL
# =========================
def get_card_synergies(card_name):
    """
    Obtiene cartas que tienen sinergia con esta carta
    Retorna dict: {nombre_carta_lower: {"synergy": score, "scryfall": url}}
    """
    url = f"https://json.edhrec.com/pages/cards/{slug(card_name)}.json"
    
    try:
        r = SESSION.get(url, timeout=10)
        if r.status_code not in [200, 403, 404]:
            return {}
        if r.status_code != 200:
            return {}
        
        data = r.json()
    except:
        return {}
    
    synergies = {}
    
    # Buscar el scryfall del objetivo (la carta que estamos analizando)
    target_scryfall = find_scryfall_in_data(data, card_name)
    
    def scan(obj):
        if isinstance(obj, dict):
            # Buscar secciones de sinergia
            if "cardviews" in obj:
                for card in obj["cardviews"]:
                    name = card.get("name")
                    synergy = card.get("synergy", 0)
                    scryfall = card.get("scryfall_uri", "")
                    
                    if name and synergy > MIN_SYNERGY_SCORE:
                        synergies[name.lower()] = {
                            "synergy": synergy,
                            "scryfall": scryfall
                        }
            
            for v in obj.values():
                scan(v)
        elif isinstance(obj, list):
            for item in obj:
                scan(item)
    
    scan(data)
    return synergies


# =========================
# 6B. BUSCAR SCRYFALL EN DATA
# =========================
def find_scryfall_in_data(data, card_name):
    """
    Busca el link de Scryfall para una carta específica en los datos de EDHREC
    """
    if data is None:
        return ""
    
    target = card_name.lower()
    found = ""
    
    def scan(obj):
        nonlocal found
        if found:
            return
        
        if isinstance(obj, dict):
            if obj.get("name") and obj["name"].lower() == target:
                if "scryfall_uri" in obj:
                    found = obj["scryfall_uri"]
                    return
            
            for v in obj.values():
                scan(v)
        elif isinstance(obj, list):
            for item in obj:
                scan(item)
    
    scan(data)
    return found


# =========================
# 6C. GENERAR URL DE SCRYFALL COMO FALLBACK
# =========================
def generate_scryfall_url(card_name):
    """
    Genera URL de Scryfall basada en el nombre de la carta
    Formato: https://scryfall.com/search?q=!"Exact Card Name"
    """
    import urllib.parse
    encoded_name = urllib.parse.quote(f'!"{card_name}"')
    return f"https://scryfall.com/search?q={encoded_name}"


# =========================
# 7. ANALIZAR Y COMPLETAR MAZO
# =========================
def analyze_and_complete_deck(deck, inventory, commander_colors):
    """
    Analiza un mazo parcial y sugiere cartas del inventario
    """
    commander = deck["commander"]
    current_cards = [c.lower() for c in deck["cards"]]
    current_cards_set = set(current_cards)
    
    print(f"\n{'='*70}")
    print(f"🔍 Analizando: {commander}")
    print(f"{'='*70}")
    print(f"   Cartas actuales: {len(current_cards)}")
    print(f"   Cartas faltantes: {64 - len(current_cards)}")
    print(f"   Colores: {', '.join(commander_colors) if commander_colors else 'Desconocidos'}")
    
    # 1. Obtener recomendaciones de EDHREC
    print(f"\n📊 Consultando EDHREC average/budget deck...")
    edhrec_cards = get_average_and_budget_deck(commander)
    time.sleep(RATE_LIMIT)
    
    # 2. Calcular sinergias con cartas actuales
    print(f"🔗 Calculando sinergias con {len(current_cards)} cartas actuales...")
    all_synergies = defaultdict(lambda: {"score": 0.0, "scryfall": ""})
    
    for idx, card in enumerate(current_cards[:20], 1):  # Limitar a 20 para no tardar mucho
        print(f"   [{idx}/20] Sinergias de: {card}")
        synergies = get_card_synergies(card)
        
        for synergy_card, data in synergies.items():
            all_synergies[synergy_card]["score"] += data["synergy"]
            if data["scryfall"] and not all_synergies[synergy_card]["scryfall"]:
                all_synergies[synergy_card]["scryfall"] = data["scryfall"]
        
        time.sleep(RATE_LIMIT)
    
    # 3. Compilar sugerencias CON VERIFICACIÓN DE COLORES
    suggestions = []
    key_cards_missing = []
    
    print(f"\n🎨 Filtrando cartas por identidad de colores...")
    print(f"   Colores del comandante: {commander_colors}")
    
    # Crear lista de cartas a verificar
    cards_to_check = []
    
    for card_lower, info in edhrec_cards.items():
        if card_lower in current_cards_set:
            continue
        if card_lower not in inventory:
            # Cartas clave faltantes (no verificar colores, solo informar)
            inclusion_pct = info.get("inclusion", 0) / info.get("num_decks", 1) if info.get("num_decks", 0) > 0 else 0
            price = info.get("price", 999)
            
            if inclusion_pct >= KEY_CARD_INCLUSION_MIN and price <= KEY_CARD_PRICE_MAX:
                scryfall_url = info.get("scryfall", "")
                if not scryfall_url:
                    scryfall_url = generate_scryfall_url(info["name"])
                
                key_cards_missing.append({
                    "name": info["name"],
                    "inclusion": inclusion_pct,
                    "price": price,
                    "num_decks": info.get("num_decks", 0),
                    "scryfall": scryfall_url
                })
            continue
        
        cards_to_check.append(("edhrec", card_lower, info))
    
    # Agregar cartas de sinergia a verificar
    for card_lower, syn_data in all_synergies.items():
        if card_lower in current_cards_set:
            continue
        if card_lower in edhrec_cards:
            continue
        if card_lower not in inventory:
            continue
        
        synergy_score = syn_data.get("score", 0)
        combined_score = min(synergy_score * 10, 50)
        
        if combined_score > MIN_SCORE_THRESHOLD:
            cards_to_check.append(("synergy", card_lower, syn_data))
    
    print(f"   Verificando {len(cards_to_check)} cartas...")
    
    # Verificar colores en lotes
    legal_count = 0
    illegal_count = 0
    
    for idx, (source_type, card_lower, data) in enumerate(cards_to_check, 1):
        if idx % 10 == 0:
            print(f"   [{idx}/{len(cards_to_check)}] Legal: {legal_count}, Ilegal: {illegal_count}")
        
        card_name = data["name"] if source_type == "edhrec" else inventory[card_lower]["name"]
        
        # Verificar colores
        if CHECK_COLOR_IDENTITY and commander_colors:
            is_legal = card_is_legal_in_colors(card_name, commander_colors, inventory)
            time.sleep(COLOR_CHECK_RATE_LIMIT)
            
            if not is_legal:
                illegal_count += 1
                continue
        
        legal_count += 1
        
        # Calcular score y agregar
        if source_type == "edhrec":
            info = data
            inclusion_pct = info.get("inclusion", 0) / info.get("num_decks", 1) if info.get("num_decks", 0) > 0 else 0
            synergy_score = all_synergies.get(card_lower, {}).get("score", 0)
            synergy_scryfall = all_synergies.get(card_lower, {}).get("scryfall", "")
            is_budget = info.get("is_budget", False)
            
            combined_score = (
                inclusion_pct * 100 +
                min(synergy_score * 10, 50) +
                (10 if is_budget else 0)
            )
            
            if combined_score < MIN_SCORE_THRESHOLD:
                continue
            
            scryfall_url = info.get("scryfall", "") or synergy_scryfall
            if not scryfall_url:
                scryfall_url = generate_scryfall_url(info["name"])
            
            suggestions.append({
                "name": info["name"],
                "score": combined_score,
                "inclusion": inclusion_pct,
                "synergy": synergy_score,
                "is_budget": is_budget,
                "source": "EDHREC",
                "collections": "; ".join(inventory[card_lower]["collections"]),
                "scryfall": scryfall_url
            })
        
        else:  # synergy
            syn_data = data
            synergy_score = syn_data.get("score", 0)
            combined_score = min(synergy_score * 10, 50)
            
            if combined_score < MIN_SCORE_THRESHOLD:
                continue
            
            scryfall_url = syn_data.get("scryfall", "")
            if not scryfall_url:
                scryfall_url = generate_scryfall_url(inventory[card_lower]["name"])
            
            suggestions.append({
                "name": inventory[card_lower]["name"],
                "score": combined_score,
                "inclusion": 0,
                "synergy": synergy_score,
                "is_budget": False,
                "source": "Sinergia",
                "collections": "; ".join(inventory[card_lower]["collections"]),
                "scryfall": scryfall_url
            })
    
    print(f"   ✅ Verificación completa: {legal_count} legales, {illegal_count} ilegales")
    
    # Ordenar sugerencias
    suggestions.sort(key=lambda x: x["score"], reverse=True)
    key_cards_missing.sort(key=lambda x: x["inclusion"], reverse=True)
    
    return {
        "commander": commander,
        "current_size": len(current_cards),
        "cards_needed": 64 - len(current_cards),
        "suggestions": suggestions[:MAX_SUGGESTIONS],
        "key_cards_missing": key_cards_missing[:20]
    }


# =========================
# 8. GENERAR REPORTE
# =========================
def generate_completion_report(analysis_results):
    """
    Genera archivos de texto y Excel con las sugerencias
    """
    import os
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    for result in analysis_results:
        commander = result["commander"]
        safe_name = slug(commander)
        
        # Archivo de texto
        txt_file = os.path.join(OUTPUT_DIR, f"{safe_name}_sugerencias.txt")
        
        with open(txt_file, 'w', encoding='utf-8') as f:
            f.write(f"# SUGERENCIAS PARA COMPLETAR: {commander}\n")
            f.write(f"# Cartas actuales: {result['current_size']}\n")
            f.write(f"# Cartas necesarias: {result['cards_needed']}\n\n")
            
            # Sugerencias
            f.write(f"## CARTAS SUGERIDAS DE TU INVENTARIO ({len(result['suggestions'])})\n")
            f.write(f"## Top {MAX_SUGGESTIONS} ordenadas por relevancia\n\n")
            
            for idx, card in enumerate(result["suggestions"], 1):
                budget_mark = " 💰" if card["is_budget"] else ""
                synergy_mark = " 🔗" if card["synergy"] > 0.3 else ""
                
                f.write(f"{idx:2d}. {card['name']}{budget_mark}{synergy_mark}\n")
                f.write(f"    Score: {card['score']:.1f} | ")
                f.write(f"Inclusión: {card['inclusion']:.1%} | ")
                f.write(f"Sinergia: {card['synergy']:.2f} | ")
                f.write(f"Fuente: {card['source']}\n")
                f.write(f"    Colecciones: {card['collections']}\n")
                if card.get("scryfall"):
                    f.write(f"    Scryfall: {card['scryfall']}\n")
                f.write("\n")
            
            # Cartas clave faltantes
            if result["key_cards_missing"]:
                f.write(f"\n## ⚠️  CARTAS CLAVE QUE NO TIENES (deberías comprar)\n")
                f.write(f"## Inclusión >40%, Precio <$2 USD\n\n")
                
                for idx, card in enumerate(result["key_cards_missing"], 1):
                    f.write(f"{idx:2d}. {card['name']}\n")
                    f.write(f"    Inclusión: {card['inclusion']:.1%} | ")
                    f.write(f"Precio: ${card['price']:.2f} | ")
                    f.write(f"En {card['num_decks']} decks\n")
                    if card.get("scryfall"):
                        f.write(f"    Scryfall: {card['scryfall']}\n")
                    f.write("\n")
            
            f.write("\n## LEYENDA\n")
            f.write("# 💰 = Carta aparece en Budget EDHREC\n")
            f.write("# 🔗 = Alta sinergia con cartas actuales del mazo\n")
            f.write("# Score = Combinación de inclusión, sinergia y budget\n")
        
        print(f"   ✅ {txt_file}")
    
    # Excel consolidado
    excel_file = os.path.join(OUTPUT_DIR, "todas_sugerencias.xlsx")
    
    excel_data = []
    for result in analysis_results:
        for card in result["suggestions"]:
            excel_data.append({
                "Comandante": result["commander"],
                "Carta": card["name"],
                "Score": card["score"],
                "Inclusión %": card["inclusion"],
                "Sinergia": card["synergy"],
                "Budget": "Sí" if card["is_budget"] else "No",
                "Fuente": card["source"],
                "Colecciones": card["collections"],
                "Scryfall": card.get("scryfall", "")
            })
    
    df_suggestions = pd.DataFrame(excel_data)
    
    # Cartas clave faltantes
    key_cards_data = []
    for result in analysis_results:
        for card in result["key_cards_missing"]:
            key_cards_data.append({
                "Comandante": result["commander"],
                "Carta": card["name"],
                "Inclusión %": card["inclusion"],
                "Precio USD": card["price"],
                "# Decks": card["num_decks"],
                "Scryfall": card.get("scryfall", "")
            })
    
    df_key_cards = pd.DataFrame(key_cards_data)
    
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        if not df_suggestions.empty:
            df_suggestions.to_excel(writer, sheet_name='Sugerencias', index=False)
        
        if not df_key_cards.empty:
            df_key_cards.to_excel(writer, sheet_name='Cartas Clave Faltantes', index=False)
    
    # Aplicar hipervínculos
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    
    wb = load_workbook(excel_file)
    
    # Hoja de sugerencias
    if 'Sugerencias' in wb.sheetnames:
        ws = wb['Sugerencias']
        
        # Colores para budget
        budget_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        high_synergy_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        hyperlinks_added = 0
        for r in range(2, ws.max_row + 1):
            # Hipervínculo a Scryfall (columna I)
            scryfall_url = ws[f'I{r}'].value
            if scryfall_url and isinstance(scryfall_url, str) and scryfall_url.startswith('http'):
                try:
                    ws[f'B{r}'].hyperlink = scryfall_url
                    ws[f'B{r}'].style = "Hyperlink"
                    hyperlinks_added += 1
                except Exception as e:
                    print(f"      ⚠️  Error agregando hyperlink en fila {r}: {e}")
            
            # Color para cartas budget (columna F)
            is_budget = ws[f'F{r}'].value
            if is_budget == "Sí":
                ws[f'F{r}'].fill = budget_fill
            
            # Color para alta sinergia (columna E)
            synergy = ws[f'E{r}'].value
            if isinstance(synergy, (int, float)) and synergy > 0.5:
                ws[f'E{r}'].fill = high_synergy_fill
        
        print(f"      📎 {hyperlinks_added} hipervínculos agregados en Sugerencias")
    
    # Hoja de cartas clave
    if 'Cartas Clave Faltantes' in wb.sheetnames:
        ws = wb['Cartas Clave Faltantes']
        
        high_inclusion_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        
        hyperlinks_added = 0
        for r in range(2, ws.max_row + 1):
            # Hipervínculo a Scryfall (columna F)
            scryfall_url = ws[f'F{r}'].value
            if scryfall_url and isinstance(scryfall_url, str) and scryfall_url.startswith('http'):
                try:
                    ws[f'B{r}'].hyperlink = scryfall_url
                    ws[f'B{r}'].style = "Hyperlink"
                    hyperlinks_added += 1
                except Exception as e:
                    print(f"      ⚠️  Error agregando hyperlink en fila {r}: {e}")
            
            # Color para alta inclusión (columna C)
            inclusion = ws[f'C{r}'].value
            if isinstance(inclusion, (int, float)) and inclusion > 0.5:
                ws[f'C{r}'].fill = high_inclusion_fill
        
        print(f"      📎 {hyperlinks_added} hipervínculos agregados en Cartas Clave")
    
    wb.save(excel_file)
    
    print(f"\n   📊 {excel_file}")


# =========================
# MAIN
# =========================
if __name__ == "__main__":
    print("="*70)
    print("🔨 COMPLETADOR DE MAZOS v1.0")
    print("="*70)
    
    # Cargar inventario
    print(f"\n📦 Cargando inventario desde: {INVENTORY_FILE}")
    inventory = load_inventory()
    print(f"   ✅ {len(inventory)} cartas únicas en inventario")
    
    # Leer mazos parciales
    print(f"\n📋 Leyendo mazos parciales desde: {PARTIAL_DECKLIST_FILE}")
    decks = parse_partial_decklists(PARTIAL_DECKLIST_FILE)
    print(f"   ✅ {len(decks)} mazos encontrados")
    
    for deck in decks:
        print(f"      • {deck['commander']}: {len(deck['cards'])} cartas")
    
    # Confirmar
    print(f"\n⚠️  Se analizarán {len(decks)} mazos")
    print(f"   Esto puede tardar {len(decks) * 2}-{len(decks) * 5} minutos")
    confirm = input("\n¿Continuar? (s/n): ").strip().lower()
    
    if confirm != 's':
        print("\n❌ Operación cancelada")
        exit(0)
    
    # Analizar cada mazo
    results = []
    
    for idx, deck in enumerate(decks, 1):
        print(f"\n{'='*70}")
        print(f"📦 MAZO {idx}/{len(decks)}")
        
        # Obtener colores del comandante
        commander_colors = get_commander_colors(deck["commander"])
        time.sleep(RATE_LIMIT)
        
        # Analizar y obtener sugerencias
        result = analyze_and_complete_deck(deck, inventory, commander_colors)
        results.append(result)
        
        print(f"\n   ✅ Análisis completado")
        print(f"      • {len(result['suggestions'])} sugerencias encontradas")
        print(f"      • {len(result['key_cards_missing'])} cartas clave faltantes")
    
    # Generar reportes
    print(f"\n{'='*70}")
    print(f"📝 Generando reportes...")
    print(f"{'='*70}")
    
    generate_completion_report(results)
    
    print(f"\n{'='*70}")
    print(f"🎉 ¡Completado!")
    print(f"{'='*70}")
    print(f"\n📁 Revisa la carpeta '{OUTPUT_DIR}/' para:")
    print(f"   • Archivos .txt individuales por comandante")
    print(f"   • Excel consolidado con todas las sugerencias")
