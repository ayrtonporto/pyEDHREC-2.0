import pandas as pd
import requests
import re
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from concurrent.futures import ThreadPoolExecutor, as_completed
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry


# =========================
# CONFIG
# =========================
INVENTORY_FILE = "inventario.csv"
RATE_LIMIT = 0.001  # Reducido de 0.1 a 0.05 (más rápido, pero ajusta si hay problemas)
MIN_MATCHES = 10
MAX_WORKERS = 5  # Número de peticiones concurrentes
REQUEST_TIMEOUT = 10
MAX_RETRIES = 3


# =========================
# SESSION CON RETRY
# =========================
def create_session():
    """Crea una sesión con reintentos automáticos para errores de red"""
    session = requests.Session()
    retry_strategy = Retry(
        total=MAX_RETRIES,
        backoff_factor=1,  # Espera 1s, 2s, 4s entre reintentos
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET"]
    )
    adapter = HTTPAdapter(max_retries=retry_strategy)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    return session


# Sesión global reutilizable
SESSION = create_session()


# =========================
# 1. LOAD INVENTORY + DEDUPE
# =========================
inv = pd.read_csv(INVENTORY_FILE)
inv["name_lower"] = inv["name"].str.lower()

inv_unique = inv.groupby("name_lower", as_index=False).agg(
    {"quantity":"sum", "name":"first"}
)

inv_map = {
    row["name_lower"]: row["quantity"]
    for _, row in inv_unique.iterrows()
}

# colecciones donde aparece cada carta
inv_source_map = (
    inv.groupby("name_lower")["source"]
    .agg(lambda xs: "; ".join(sorted(set(map(str, xs)))))
    .to_dict()
)

BASIC_LANDS = {"plains","island","swamp","mountain","forest","wastes"}


# =========================
# 2. SLUG
# =========================
def slug(s: str) -> str:
    """
    Convierte nombres de cartas a formato slug para URLs de EDHREC
    Maneja tildes, acentos y caracteres especiales
    """
    import unicodedata
    
    # Normalizar unicode y remover acentos/tildes
    s = unicodedata.normalize('NFD', s)
    s = ''.join(char for char in s if unicodedata.category(char) != 'Mn')
    
    s = s.lower()
    # Remover caracteres especiales comunes
    s = s.replace("'","").replace("'","").replace("â€™","")
    s = s.replace(",","").replace(":","").replace(".","")
    # Convertir espacios y otros caracteres no alfanuméricos a guiones
    s = re.sub(r"[^a-z0-9]+","-",s)
    return s.strip("-")


# =========================
# 3. FETCH CARD JSON (CON MANEJO DE ERRORES)
# =========================
def fetch_card_json(card_name):
    url = f"https://json.edhrec.com/pages/cards/{slug(card_name)}.json"
    try:
        r = SESSION.get(url, timeout=REQUEST_TIMEOUT)
        if r.status_code == 403:
            # 403 generalmente significa que la carta no existe en EDHREC
            # (común con cartas nuevas o de sets especiales)
            return None
        if r.status_code == 404:
            # Carta no encontrada en EDHREC
            return None
        if r.status_code != 200:
            print(f"  ⚠️  Error {r.status_code} para {card_name}")
            return None
        return r.json()
    except requests.exceptions.ConnectionError:
        print(f"  ❌ Error de conexión para {card_name} - verifica tu internet")
        return None
    except requests.exceptions.Timeout:
        print(f"  ⏱️  Timeout para {card_name}")
        return None
    except Exception as e:
        print(f"  ❌ Error inesperado para {card_name}: {e}")
        return None


# =========================
# 4. FIND SCRYFALL FOR THIS CARD
# =========================
def find_scryfall_for_card(data, target_name):
    if data is None:
        return None
    target = target_name.lower()
    found = None

    def scan(o):
        nonlocal found
        if found:
            return
        if isinstance(o, dict):
            if o.get("name") and o["name"].lower() == target:
                if "scryfall_uri" in o:
                    found = o["scryfall_uri"]
                    return
            for v in o.values():
                scan(v)
        elif isinstance(o, list):
            for e in o:
                scan(e)

    scan(data)
    return found


# =========================
# 5. EXTRACT COMMANDERS (TOP COMMANDER ANALYSIS)
# =========================
def extract_commanders_from_card_json(data, card_name):
    commanders = []
    scry = find_scryfall_for_card(data, card_name)

    def scan(o):
        if isinstance(o, dict):
            url = o.get("url")
            if isinstance(url,str) and "/commanders/" in url:
                name = o.get("name")
                inc  = o.get("inclusion")
                pot  = o.get("potential_decks")
                if name and inc and pot:
                    percent = inc/pot
                    if percent > 0.2:
                        commanders.append({
                            "commander": name,
                            "percent": percent,
                            "scryfall": scry
                        })
            for v in o.values():
                scan(v)
        elif isinstance(o, list):
            for e in o:
                scan(e)

    scan(data)
    return commanders


# =========================
# 6. EXTRACT SYNERGY CARDS
# =========================
def extract_synergy_cards(data, card_name):
    """
    Extrae cartas con alta sinergia desde la sección de "Cards that work well with X"
    Esto captura sinergias que el análisis de comandantes podría perder
    """
    synergy_cards = []
    scry = find_scryfall_for_card(data, card_name)
    
    def scan(o):
        if isinstance(o, dict):
            # Buscar secciones de synergy
            if "synergy" in str(o.get("header", "")).lower() or "cards" in str(o.get("tag", "")).lower():
                if "cardviews" in o:
                    for cv in o["cardviews"]:
                        nm = cv.get("name")
                        syn = cv.get("synergy", 0)
                        if nm and syn > 0.1:  # Solo sinergias significativas
                            synergy_cards.append({
                                "card": nm,
                                "synergy": syn,
                                "with_card": card_name
                            })
            
            for v in o.values():
                scan(v)
        elif isinstance(o, list):
            for e in o:
                scan(e)
    
    scan(data)
    return synergy_cards


# =========================
# 7. FETCH AVERAGE DECK (CON MANEJO DE ERRORES)
# =========================
def fetch_average_deck(commander_name):
    url = f"https://json.edhrec.com/pages/average-decks/{slug(commander_name)}.json"
    try:
        r = SESSION.get(url, timeout=REQUEST_TIMEOUT)
        if r.status_code != 200:
            return []
        data = r.json()
    except requests.exceptions.ConnectionError:
        print(f"  ❌ Error de conexión al buscar average deck de {commander_name}")
        return []
    except requests.exceptions.Timeout:
        print(f"  ⏱️  Timeout al buscar average deck de {commander_name}")
        return []
    except Exception as e:
        print(f"  ❌ Error en average deck de {commander_name}: {e}")
        return []

    cards = set()

    def scan(o):
        if isinstance(o, dict):
            if "cardviews" in o:
                for cv in o["cardviews"]:
                    nm = cv.get("name")
                    if nm:
                        cards.add(nm)
            for v in o.values():
                scan(v)
        elif isinstance(o, list):
            for e in o:
                scan(e)

    scan(data)
    return sorted(cards)


# =========================
# 8. PROCESS SINGLE CARD
# =========================
def process_card(idx, total, card, card_lower):
    """Procesa una carta individual (para paralelización)"""
    if card_lower in BASIC_LANDS:
        return None
    
    print(f"[{idx+1}/{total}] {card}...")
    
    data = fetch_card_json(card)
    if data is None:
        time.sleep(RATE_LIMIT)
        return None
    
    coms = extract_commanders_from_card_json(data, card)
    synergies = extract_synergy_cards(data, card)
    
    time.sleep(RATE_LIMIT)
    
    return {
        "card": card,
        "commanders": coms,
        "synergies": synergies
    }


# =========================
# 9. MAIN ANALYSIS (MEJORADO)
# =========================
def analyze_inventory():
    commander_stats = {}
    synergy_recommendations = []
    skipped_cards = []  # Para trackear cartas que no se pudieron procesar
    total = len(inv_unique)
    
    print(f"\n🔍 Analizando {total} cartas únicas...")
    
    # Procesar cartas con manejo de errores robusto
    for idx, row in inv_unique.iterrows():
        card = row["name"]
        card_lower = row["name_lower"]
        
        result = process_card(idx, total, card, card_lower)
        
        if result is None:
            skipped_cards.append(card)
            continue
        
        # Procesar comandantes
        for c in result["commanders"]:
            cname = c["commander"]
            percent = c["percent"]
            scry = c["scryfall"]

            entry = commander_stats.setdefault(
                cname,
                {
                    "matches":0,
                    "percent_sum":0.0,
                    "percent_count":0,
                    "cards":[]
                }
            )

            entry["cards"].append({
                "name": card,
                "percent": percent,
                "source": "edhrec",
                "scryfall": scry
            })

            entry["matches"] += 1
            entry["percent_sum"] += percent
            entry["percent_count"] += 1
        
        # Procesar sinergias
        for syn in result["synergies"]:
            syn_card_lower = syn["card"].lower()
            if syn_card_lower in inv_map:
                synergy_recommendations.append({
                    "card_owned": card,
                    "synergy_card": syn["card"],
                    "synergy_score": syn["synergy"],
                    "both_owned": True
                })

    rows=[]
    for cname,info in commander_stats.items():
        if info["matches"] < MIN_MATCHES:
            continue

        avg = info["percent_sum"]/info["percent_count"] if info["percent_count"] else None

        rows.append({
            "commander": cname,
            "matches": info["matches"],
            "avg_percent": avg,
            "cards": info["cards"]
        })

    df = pd.DataFrame(rows)
    df = df.sort_values(by=["matches","avg_percent"], ascending=[False,False])
    
    # Informar sobre cartas omitidas
    if skipped_cards:
        print(f"\n⚠️  {len(skipped_cards)} cartas no se pudieron procesar en EDHREC:")
        print(f"    (Pueden ser cartas muy nuevas, de sets especiales, o Universe Beyond)")
        if len(skipped_cards) <= 10:
            for sc in skipped_cards:
                print(f"    - {sc}")
        else:
            for sc in skipped_cards[:10]:
                print(f"    - {sc}")
            print(f"    ... y {len(skipped_cards) - 10} más")
    
    return df, synergy_recommendations


# =========================
# 10. ANÁLISIS POR TEMAS
# =========================
def analyze_themes(flat_cards, synergies):
    """
    Detecta temas/arquetipos basándose en las cartas y comandantes
    """
    theme_keywords = {
        "Tokens": ["token", "create", "populate", "doubling season", "anointed procession"],
        "Sacrifice": ["sacrifice", "aristocrats", "blood artist", "zulaport", "mayhem devil"],
        "+1/+1 Counters": ["counter", "+1/+1", "proliferate", "modular", "evolve"],
        "Graveyard": ["graveyard", "reanimate", "flashback", "delve", "escape", "dredge"],
        "Artifacts": ["artifact", "affinity", "metalcraft", "improvise", "treasure"],
        "Enchantments": ["enchantment", "constellation", "enchantress", "saga"],
        "Spellslinger": ["instant", "sorcery", "prowess", "storm", "magecraft"],
        "Voltron": ["equipment", "aura", "voltron", "commander damage"],
        "Ramp": ["ramp", "land", "mana", "cultivate", "kodama", "explosive vegetation"],
        "Card Draw": ["draw", "card advantage", "rhystic", "curiosity", "wheel"],
        "Tribal": ["elf", "goblin", "zombie", "dragon", "changeling", "tribal"],
        "Control": ["counter", "removal", "board wipe", "control", "cyclonic rift"],
        "Combo": ["infinite", "combo", "win condition", "thoracle"],
        "Landfall": ["landfall", "land enters", "fetch", "evolving wilds"],
        "Blink": ["blink", "flicker", "enters the battlefield", "etb"],
    }
    
    theme_scores = {theme: 0 for theme in theme_keywords}
    theme_cards = {theme: [] for theme in theme_keywords}
    
    # Analizar cartas
    for card_entry in flat_cards:
        card_name = card_entry["card"].lower()
        
        for theme, keywords in theme_keywords.items():
            if any(kw in card_name for kw in keywords):
                theme_scores[theme] += 1
                if len(theme_cards[theme]) < 10:  # Limitar a 10 ejemplos
                    theme_cards[theme].append(card_entry["card"])
    
    # Crear DataFrame
    theme_rows = []
    for theme, score in theme_scores.items():
        if score >= 3:  # Solo temas con al menos 3 cartas
            theme_rows.append({
                "Tema/Arquetipo": theme,
                "Cartas Detectadas": score,
                "Ejemplos": ", ".join(theme_cards[theme][:5]),
                "Potencial": "Alto" if score >= 10 else "Medio" if score >= 6 else "Bajo"
            })
    
    df_themes = pd.DataFrame(theme_rows)
    if not df_themes.empty:
        df_themes = df_themes.sort_values(by="Cartas Detectadas", ascending=False)
    
    return df_themes


# =========================
# 11. RESUMEN EJECUTIVO
# =========================
def create_executive_summary(commanders_df, cards_df, synergies):
    """
    Crea un resumen con las estadísticas clave del análisis
    """
    summary_data = []
    
    # Estadísticas generales
    summary_data.append({
        "Métrica": "Total de Comandantes Viables",
        "Valor": len(commanders_df),
        "Descripción": f"Comandantes con al menos {MIN_MATCHES} cartas en tu colección"
    })
    
    summary_data.append({
        "Métrica": "Total de Cartas Únicas Analizadas",
        "Valor": len(inv_unique),
        "Descripción": "Cartas únicas en tu inventario (excluyendo tierras básicas)"
    })
    
    if not commanders_df.empty:
        top_commander = commanders_df.iloc[0]
        summary_data.append({
            "Métrica": "Mejor Comandante Recomendado",
            "Valor": top_commander["commander"],
            "Descripción": f"{top_commander['matches']} cartas coincidentes, {top_commander['avg_percent']:.1%} promedio inclusión"
        })
    
    # Conteo por fuente
    both_count = len(cards_df[cards_df["source"] == "both"])
    edhrec_count = len(cards_df[cards_df["source"] == "edhrec"])
    average_count = len(cards_df[cards_df["source"] == "average"])
    
    summary_data.append({
        "Métrica": "Cartas con Doble Validación (BOTH)",
        "Valor": both_count,
        "Descripción": "Las más confiables - aparecen en Top Commanders Y Average Deck"
    })
    
    summary_data.append({
        "Métrica": "Cartas Solo en Top Commanders",
        "Valor": edhrec_count,
        "Descripción": "Populares en análisis de comandantes"
    })
    
    summary_data.append({
        "Métrica": "Cartas Solo en Average Deck",
        "Valor": average_count,
        "Descripción": "Staples del mazo promedio"
    })
    
    if synergies:
        summary_data.append({
            "Métrica": "Sinergias Detectadas",
            "Valor": len(synergies),
            "Descripción": "Pares de cartas con alta sinergia mutua"
        })
    
    # Top 5 comandantes
    summary_data.append({
        "Métrica": "--- TOP 5 COMANDANTES ---",
        "Valor": "",
        "Descripción": ""
    })
    
    for idx, row in commanders_df.head(5).iterrows():
        summary_data.append({
            "Métrica": f"#{idx+1}",
            "Valor": row["commander"],
            "Descripción": f"{row['matches']} cartas | {row['avg_percent']:.1%} avg"
        })
    
    return pd.DataFrame(summary_data)


# =========================
# 12. EXPORT CON MÚLTIPLES HOJAS
# =========================
def export_with_formatting(df, synergies):
    flat = []

    print("\n📊 Generando reporte detallado...")
    
    # ---- 1. EDHREC CARDS ----
    for _, row in df.iterrows():
        commander = row["commander"]
        for c in row["cards"]:
            card = c["name"]
            lower = card.lower()

            flat.append({
                "commander": commander,
                "card": card,
                "percent": c["percent"],
                "scryfall": c["scryfall"],
                "source": "edhrec",
                "collections": inv_source_map.get(lower, "")
            })

    # ---- 2. AVERAGE DECK CARDS ----
    total_commanders = len(df)
    for idx, row in df.iterrows():
        commander = row["commander"]
        print(f"  [{idx+1}/{total_commanders}] Consultando average deck de {commander}...")
        
        avg_cards = fetch_average_deck(commander)

        for card in avg_cards:
            lower = card.lower()
            if lower not in inv_map:
                continue

            data = fetch_card_json(card)
            scry = find_scryfall_for_card(data, card)

            collections = inv_source_map.get(lower, "")

            # check if already added
            already = next((x for x in flat if x["commander"]==commander and x["card"].lower()==lower), None)

            if already:
                already["source"] = "both"
            else:
                flat.append({
                    "commander": commander,
                    "card": card,
                    "percent": None,
                    "scryfall": scry,
                    "source": "average",
                    "collections": collections
                })
            
            time.sleep(RATE_LIMIT)

    df2 = pd.DataFrame(flat)

    # ---- ORDENAMIENTO POR CANTIDAD DE CARTAS + PRIORIDAD ----
    def get_sort_priority(row):
        source = row["source"]
        percent = row["percent"] if pd.notna(row["percent"]) else 0
        
        if source == "both":
            return (1, -percent)  # Amarillo primero, ordenado por percent desc
        elif source == "edhrec" and percent > 0.20:
            return (2, -percent)  # Verde segundo
        elif source == "average":
            return (3, 0)  # Azul tercero
        else:  # edhrec con percent < 0.20
            return (4, -percent)  # Rojo último
    
    # Calcular total de cartas por comandante para ordenar
    commander_card_count = df2.groupby("commander").size().to_dict()
    df2["commander_total"] = df2["commander"].map(commander_card_count)
    
    df2["sort_key"] = df2.apply(lambda row: get_sort_priority(row), axis=1)
    df2 = df2.sort_values(
        by=["commander_total", "commander", "sort_key"],
        ascending=[False, True, True]
    )
    df2 = df2.drop(columns=["sort_key", "commander_total"])

    out = "commander_matches_formatted.xlsx"
    
    # ---- MANEJO DE ARCHIVO BLOQUEADO ----
    # Si el archivo está abierto en Excel, intentar con nombre alternativo
    try:
        # Intentar eliminar archivo antiguo si existe
        import os
        if os.path.exists(out):
            try:
                os.remove(out)
            except PermissionError:
                # Archivo abierto, usar nombre alternativo
                import datetime
                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                out = f"commander_matches_{timestamp}.xlsx"
                print(f"\n⚠️  El archivo original está abierto. Guardando como: {out}")
    except Exception as e:
        print(f"⚠️  Advertencia al verificar archivo: {e}")
    
    # ---- CREAR ARCHIVO EXCEL CON MÚLTIPLES HOJAS ----
    with pd.ExcelWriter(out, engine='openpyxl') as writer:
        df2.to_excel(writer, sheet_name='Comandantes Recomendados', index=False)
        
        # ---- SEGUNDA HOJA: ANÁLISIS DE SINERGIAS ----
        if synergies:
            synergy_df = pd.DataFrame(synergies)
            
            # Agregar información de colecciones
            synergy_df["card_owned_collections"] = synergy_df["card_owned"].str.lower().map(inv_source_map)
            synergy_df["synergy_card_collections"] = synergy_df["synergy_card"].str.lower().map(inv_source_map)
            
            # Ordenar por synergy score descendente
            synergy_df = synergy_df.sort_values(by="synergy_score", ascending=False)
            
            synergy_df.to_excel(writer, sheet_name='Sinergias Detectadas', index=False)
        
        # ---- TERCERA HOJA: ANÁLISIS POR TEMAS ----
        theme_analysis = analyze_themes(flat, synergies)
        if not theme_analysis.empty:
            theme_analysis.to_excel(writer, sheet_name='Análisis por Temas', index=False)
        
        # ---- CUARTA HOJA: RESUMEN EJECUTIVO ----
        summary = create_executive_summary(df, df2, synergies)
        summary.to_excel(writer, sheet_name='Resumen Ejecutivo', index=False)

    # ---- FORMATO CON COLORES EN PRIMERA HOJA ----
    wb = load_workbook(out)
    ws = wb['Comandantes Recomendados']

    green = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    red   = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    blue  = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    yellow= PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    for r in range(2, ws.max_row+1):
        percent = ws[f"C{r}"].value
        source  = ws[f"E{r}"].value

        if source == "both":
            ws[f"C{r}"].fill = yellow
        elif source == "average":
            ws[f"C{r}"].fill = blue
        elif isinstance(percent,(int,float)) and percent > 0.5:
            ws[f"C{r}"].fill = green
        elif isinstance(percent,(int,float)) and percent > 0.20:
            ws[f"C{r}"].fill = red

        url = ws[f"D{r}"].value
        if url:
            ws[f"B{r}"].hyperlink = url
            ws[f"B{r}"].style = "Hyperlink"
    
    # ---- FORMATO EN HOJA DE SINERGIAS ----
    if 'Sinergias Detectadas' in wb.sheetnames:
        ws_syn = wb['Sinergias Detectadas']
        high_syn = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        med_syn = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        
        for r in range(2, ws_syn.max_row+1):
            synergy_val = ws_syn[f"C{r}"].value
            if isinstance(synergy_val, (int, float)):
                if synergy_val > 0.3:
                    ws_syn[f"C{r}"].fill = high_syn
                elif synergy_val > 0.15:
                    ws_syn[f"C{r}"].fill = med_syn

    wb.save(out)
    print(f"\n✅ Archivo generado: {out}")
    print(f"\n📋 Contenido del archivo:")
    print(f"  📄 Hoja 1: Comandantes Recomendados (ordenados por cantidad de cartas)")
    if synergies:
        print(f"  🔗 Hoja 2: Sinergias Detectadas entre tus cartas ({len(synergies)} encontradas)")
    print(f"  🎨 Hoja 3: Análisis por Temas/Arquetipos")
    print(f"  📊 Hoja 4: Resumen Ejecutivo")
    print(f"\n📋 Leyenda de colores (Hoja 1):")
    print(f"  🟡 Amarillo (BOTH): Validado en Top Commanders Y Average Deck - MÁS IMPORTANTE")
    print(f"  🟢 Verde: >20% inclusión en Top Commanders")
    print(f"  🔵 Azul: Solo en Average Deck")
    print(f"  🔴 Rojo: 5-20% inclusión en Top Commanders")
    if synergies:
        print(f"\n📋 Leyenda de colores (Hoja 2 - Sinergias):")
        print(f"  🟢 Verde claro: Sinergia alta (>30%)")
        print(f"  🟡 Amarillo: Sinergia media (15-30%)")
    
    # Retornar el DataFrame de cartas para uso posterior
    return df2


# =========================
# MAIN
# =========================
if __name__ == "__main__":
    print("="*60)
    print("🎮 ANALIZADOR DE COMANDANTES MEJORADO v2.0")
    print("="*60)
    print(f"⚙️  Configuración:")
    print(f"   - Rate limit: {RATE_LIMIT}s")
    print(f"   - Mínimo de coincidencias: {MIN_MATCHES}")
    print(f"   - Reintentos automáticos: {MAX_RETRIES}")
    print("="*60)
    
    try:
        df, synergies = analyze_inventory()

        if df.empty:
            print("\n⚠️  No hubo comandantes con suficientes coincidencias.")
            print("   Intenta reducir MIN_MATCHES en la configuración.")
        else:
            print("\n" + "="*60)
            print("🏆 TOP 15 COMANDANTES RECOMENDADOS")
            print("="*60)
            print(df[["commander","matches","avg_percent"]].head(15).to_string(index=False))
            
            # Exportar análisis completo
            cards_df = export_with_formatting(df, synergies)
            
            # Nueva funcionalidad: Selección de comandantes para deckbuilding
            print("\n" + "="*60)
            print("🎯 GENERADOR DE DECKLISTS")
            print("="*60)
            
            response = input("\n¿Deseas generar decklists para comandantes específicos? (s/n): ").strip().lower()
            
            if response == 's':
                generate_decklists_interactive(df, cards_df)
            
    except KeyboardInterrupt:
        print("\n\n⚠️  Proceso interrumpido por el usuario")
        print("   Los resultados parciales NO se guardaron")
    except Exception as e:
        print(f"\n\n❌ Error fatal: {e}")
        print("   Verifica tu conexión a internet y el formato de inventario.csv")
        import traceback
        traceback.print_exc()


# =========================
# 13. SELECCIÓN INTERACTIVA DE COMANDANTES
# =========================
def generate_decklists_interactive(commanders_df, cards_df):
    """
    Permite al usuario seleccionar comandantes y genera decklists con
    las cartas organizadas por colección
    """
    print("\n📋 Comandantes disponibles:")
    print("-" * 60)
    
    # Mostrar lista numerada de comandantes
    commander_list = commanders_df["commander"].tolist()
    for idx, commander in enumerate(commander_list, 1):
        matches = commanders_df[commanders_df["commander"] == commander]["matches"].iloc[0]
        avg_pct = commanders_df[commanders_df["commander"] == commander]["avg_percent"].iloc[0]
        print(f"{idx:2d}. {commander:40s} ({matches} cartas, {avg_pct:.1%} avg)")
    
    print("-" * 60)
    print("\nIngresa los números de los comandantes que quieres construir")
    print("Ejemplos: '1,3,5' o '1-5' o '1,3-6,10'")
    print("Escribe 'all' para seleccionar todos")
    
    selection = input("\nTu selección: ").strip()
    
    # Parsear selección
    selected_indices = parse_selection(selection, len(commander_list))
    
    if not selected_indices:
        print("\n❌ Selección inválida. Abortando.")
        return
    
    selected_commanders = [commander_list[i] for i in selected_indices]
    
    print(f"\n✅ Seleccionados {len(selected_commanders)} comandantes:")
    for cmd in selected_commanders:
        print(f"   - {cmd}")
    
    # Generar decklists
    print("\n🔨 Generando decklists...")
    generate_decklists(selected_commanders, cards_df, commanders_df)


# =========================
# 14. PARSEAR SELECCIÓN DE USUARIO
# =========================
def parse_selection(selection, max_index):
    """
    Parsea input del usuario como '1,3,5' o '1-5' o 'all'
    Retorna lista de índices (0-based)
    """
    if selection.lower() == 'all':
        return list(range(max_index))
    
    indices = set()
    
    try:
        parts = selection.split(',')
        for part in parts:
            part = part.strip()
            if '-' in part:
                # Rango: '1-5'
                start, end = part.split('-')
                start = int(start.strip()) - 1  # Convert to 0-based
                end = int(end.strip()) - 1
                if start < 0 or end >= max_index or start > end:
                    return []
                indices.update(range(start, end + 1))
            else:
                # Número individual: '3'
                num = int(part.strip()) - 1  # Convert to 0-based
                if num < 0 or num >= max_index:
                    return []
                indices.add(num)
        
        return sorted(list(indices))
    except:
        return []


# =========================
# 15. GENERAR DECKLISTS POR COMANDANTE
# =========================
def generate_decklists(selected_commanders, cards_df, commanders_df):
    """
    Genera una única decklist consolidada con todas las cartas necesarias
    para los comandantes seleccionados, organizadas por colección.
    Cada carta tiene el tag #nombre_del_comandante al final.
    """
    import os
    
    # Crear carpeta para decklists
    output_dir = "decklists"
    os.makedirs(output_dir, exist_ok=True)
    
    print(f"\n📝 Generando decklist consolidada para {len(selected_commanders)} comandantes...")
    
    # Recopilar TODAS las cartas de los comandantes seleccionados
    all_cards = []
    
    for commander in selected_commanders:
        commander_cards = cards_df[cards_df["commander"] == commander].copy()
        
        for _, card in commander_cards.iterrows():
            all_cards.append({
                "card_name": card["card"],
                "commander": commander,
                "collection": card["collections"] if pd.notna(card["collections"]) else "Sin colección",
                "source": card["source"],
                "percent": card["percent"] if pd.notna(card["percent"]) else 0,
                "scryfall": card["scryfall"] if pd.notna(card["scryfall"]) else ""
            })
    
    # Crear DataFrame consolidado
    df_consolidated = pd.DataFrame(all_cards)
    
    # Agrupar cartas duplicadas (misma carta para múltiples comandantes)
    df_grouped = df_consolidated.groupby("card_name").agg({
        "commander": lambda x: list(x),
        "collection": "first",  # Asumimos que la colección es la misma
        "source": "first",
        "percent": "first",
        "scryfall": "first"
    }).reset_index()
    
    # Crear tags de comandantes
    df_grouped["commander_tags"] = df_grouped["commander"].apply(
        lambda cmds: " ".join([f"#{sanitize_commander_name(cmd)}" for cmd in cmds])
    )
    
    # Nombre del archivo
    if len(selected_commanders) == 1:
        safe_name = sanitize_commander_name(selected_commanders[0])
        filename_base = f"{safe_name}_decklist"
    else:
        filename_base = f"consolidated_{len(selected_commanders)}_commanders"
    
    # Generar archivo de texto
    txt_file = os.path.join(output_dir, f"{filename_base}.txt")
    generate_consolidated_text_decklist(df_grouped, selected_commanders, txt_file)
    
    # Generar Excel
    xlsx_file = os.path.join(output_dir, f"{filename_base}.xlsx")
    generate_consolidated_excel(df_grouped, selected_commanders, xlsx_file)
    
    print(f"\n✅ Decklist consolidada generada:")
    print(f"   📄 {txt_file}")
    print(f"   📊 {xlsx_file}")
    print(f"\n📋 Total de cartas únicas a extraer: {len(df_grouped)}")
    
    # Mostrar resumen por colección
    print(f"\n📦 Resumen por colección:")
    collection_summary = df_grouped.groupby("collection").size().sort_values(ascending=False)
    for collection, count in collection_summary.items():
        print(f"   {collection:30s} {count:3d} cartas")


# =========================
# 16. SANITIZAR NOMBRE DE COMANDANTE
# =========================
def sanitize_commander_name(commander):
    """
    Convierte nombre de comandante a formato tag limpio
    Ejemplo: "Atraxa, Praetors' Voice" -> "Atraxa_Praetors_Voice"
    """
    # Remover caracteres especiales
    clean = commander.replace(",", "").replace("'", "").replace(":", "")
    clean = clean.replace("/", "_").replace("-", "_")
    # Reemplazar espacios con guiones bajos
    clean = "_".join(clean.split())
    return clean


# =========================
# 17. GENERAR DECKLIST TEXTO CONSOLIDADA
# =========================
def generate_consolidated_text_decklist(df_grouped, commanders, output_file):
    """
    Genera archivo .txt con todas las cartas organizadas por colección
    Cada carta tiene #comandante al final
    """
    with open(output_file, 'w', encoding='utf-8') as f:
        # Encabezado
        f.write(f"# DECKLIST CONSOLIDADA\n")
        f.write(f"# Comandantes: {', '.join(commanders)}\n")
        f.write(f"# Total de cartas únicas: {len(df_grouped)}\n")
        f.write(f"# Generado por Analizador de Comandantes v2.0\n\n")
        
        # Listar comandantes
        f.write("## COMANDANTES\n")
        for cmd in commanders:
            f.write(f"1 {cmd}\n")
        f.write("\n")
        
        # Agrupar por colección
        collections = df_grouped["collection"].unique()
        
        for collection in sorted(collections):
            collection_cards = df_grouped[df_grouped["collection"] == collection].copy()
            
            if len(collection_cards) == 0:
                continue
            
            f.write(f"## {collection} ({len(collection_cards)} cartas)\n")
            
            # Ordenar por prioridad
            priority_order = {"both": 1, "edhrec": 2, "average": 3}
            collection_cards["priority"] = collection_cards["source"].map(priority_order).fillna(4)
            collection_cards = collection_cards.sort_values(
                by=["priority", "percent"],
                ascending=[True, False]
            )
            
            for _, card in collection_cards.iterrows():
                # Marcador de prioridad
                source_mark = ""
                if card["source"] == "both":
                    source_mark = " ⭐"
                elif card["source"] == "edhrec" and card["percent"] > 0.20:
                    source_mark = " 🔥"
                
                # Escribir carta con tags
                f.write(f"1 {card['card_name']}{source_mark} {card['commander_tags']}\n")
            
            f.write("\n")
        
        # Leyenda
        f.write("## LEYENDA\n")
        f.write("# ⭐ = Validación doble (Top Commanders + Average Deck)\n")
        f.write("# 🔥 = Alta inclusión (>20% en Top Commanders)\n")
        f.write("# #Comandante = Tag indicando para qué deck es la carta\n")
        f.write("#\n")
        f.write("# Si una carta tiene múltiples tags, significa que se usa en varios decks\n")


# =========================
# 18. GENERAR EXCEL CONSOLIDADA
# =========================
def generate_consolidated_excel(df_grouped, commanders, output_file):
    """
    Genera archivo Excel con vista consolidada de cartas
    """
    # Preparar datos para Excel
    excel_data = []
    
    for _, card in df_grouped.iterrows():
        num_decks = len(card["commander"])
        commander_list = ", ".join(card["commander"])
        
        excel_data.append({
            "Carta": card["card_name"],
            "Colección": card["collection"],
            "Comandantes": commander_list,
            "# Decks": num_decks,
            "Fuente": card["source"],
            "Inclusión %": card["percent"] if card["percent"] > 0 else None,
            "Prioridad": "⭐ Alta" if card["source"] == "both" else 
                        ("🔥 Media-Alta" if (card["source"] == "edhrec" and card["percent"] > 0.20) else
                         ("📘 Media" if card["source"] == "average" else "📗 Baja")),
            "Scryfall": card["scryfall"],
            "Tags": card["commander_tags"]
        })
    
    df_excel = pd.DataFrame(excel_data)
    
    # Ordenar por colección y prioridad
    priority_map = {"⭐ Alta": 1, "🔥 Media-Alta": 2, "📘 Media": 3, "📗 Baja": 4}
    df_excel["_sort"] = df_excel["Prioridad"].map(priority_map)
    df_excel = df_excel.sort_values(
        by=["Colección", "_sort", "# Decks", "Inclusión %"],
        ascending=[True, True, False, False]
    )
    df_excel = df_excel.drop(columns=["_sort"])
    
    # Crear Excel con múltiples hojas
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Hoja 1: Lista completa por colección
        df_excel.to_excel(writer, sheet_name='Por Colección', index=False)
        
        # Hoja 2: Resumen por colección
        summary = df_excel.groupby("Colección").agg({
            "Carta": "count",
            "# Decks": "sum",
            "Prioridad": lambda x: f"{sum(x == '⭐ Alta')} altas, {sum(x.str.contains('Media'))} medias"
        }).reset_index()
        summary.columns = ["Colección", "Total Cartas", "Total Usos", "Distribución"]
        summary = summary.sort_values(by="Total Cartas", ascending=False)
        summary.to_excel(writer, sheet_name='Resumen por Colección', index=False)
        
        # Hoja 3: Cartas compartidas entre decks
        shared_cards = df_excel[df_excel["# Decks"] > 1].copy()
        if not shared_cards.empty:
            shared_cards = shared_cards.sort_values(by="# Decks", ascending=False)
            shared_cards.to_excel(writer, sheet_name='Cartas Compartidas', index=False)
        
        # Hoja 4: Lista simple para imprimir
        simple = df_excel[["Carta", "Colección", "Tags"]].copy()
        simple.to_excel(writer, sheet_name='Lista para Imprimir', index=False)
        
        # Hoja 5: Por comandante individual
        for commander in commanders:
            safe_name = sanitize_commander_name(commander)[:31]  # Excel limit
            cmd_cards = df_excel[df_excel["Comandantes"].str.contains(commander, regex=False)].copy()
            if not cmd_cards.empty:
                cmd_cards.to_excel(writer, sheet_name=safe_name, index=False)
    
    # Aplicar formato con colores
    wb = load_workbook(output_file)
    ws = wb['Por Colección']
    
    # Colores
    high = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    med_high = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    med = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    low = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    shared = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
    
    for r in range(2, ws.max_row + 1):
        priority = ws[f"G{r}"].value
        num_decks = ws[f"D{r}"].value
        
        # Resaltar cartas compartidas
        if isinstance(num_decks, int) and num_decks > 1:
            ws[f"D{r}"].fill = shared
        
        # Color por prioridad
        if priority and "⭐" in priority:
            ws[f"G{r}"].fill = high
        elif priority and "🔥" in priority:
            ws[f"G{r}"].fill = med_high
        elif priority and "📘" in priority:
            ws[f"G{r}"].fill = med
        elif priority and "📗" in priority:
            ws[f"G{r}"].fill = low
        
        # Hipervínculos
        url = ws[f"H{r}"].value
        if url:
            ws[f"A{r}"].hyperlink = url
            ws[f"A{r}"].style = "Hyperlink"
    
    wb.save(output_file)


# =========================
# 18. MODIFICAR EXPORT PARA RETORNAR DATOS
# =========================
